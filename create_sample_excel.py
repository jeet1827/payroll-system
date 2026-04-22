import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import random

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Employee Payroll"

headers = [
    "Employee ID", "Employee Name", "Department", "Designation", "Month", "Year",
    "Working Days", "Days Present", "Basic Salary", "HRA", "Conveyance Allowance",
    "Medical Allowance", "Special Allowance", "Bonus", "Overtime Pay",
    "PF Employee", "PF Employer", "ESI Employee", "ESI Employer",
    "Professional Tax", "TDS", "Loan Deduction", "Advance Deduction",
    "Leave Deduction", "Gratuity", "Bank Account", "IFSC Code", "PAN Number",
    "UAN Number", "Date of Joining"
]

header_fill = PatternFill("solid", start_color="1a3c5e", end_color="1a3c5e")
header_font = Font(bold=True, color="FFFFFF", size=11)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[cell.column_letter].width = 18

ws.row_dimensions[1].height = 35

departments = ["Engineering", "Sales", "HR", "Finance", "Operations", "Marketing", "IT Support", "Legal"]
designations = ["Manager", "Senior Engineer", "Analyst", "Executive", "Officer", "Specialist", "Lead", "Associate"]
names = [
    "Arjun Sharma", "Priya Patel", "Rahul Kumar", "Sneha Gupta", "Vikram Singh",
    "Ananya Roy", "Rohan Mehta", "Kavya Nair", "Aditya Joshi", "Pooja Iyer",
    "Suresh Reddy", "Deepika Rao", "Nikhil Verma", "Anjali Mishra", "Sanjay Das",
    "Meera Pillai", "Amit Choudhary", "Ritu Agarwal", "Kunal Bose", "Swati Shah",
    "Manish Tiwari", "Divya Kapoor", "Rajesh Pandey", "Neha Bhatt", "Sachin Yadav",
    "Lakshmi Nambiar", "Vivek Sinha", "Preeti Dubey", "Harsh Malhotra", "Sunita Patil"
]

for i, name in enumerate(names, 1):
    emp_id = f"EMP{1000+i:04d}"
    dept = departments[i % len(departments)]
    desig = designations[i % len(designations)]
    basic = random.randint(20000, 80000)
    hra = round(basic * 0.40)
    conveyance = 1600
    medical = 1250
    special = random.randint(2000, 8000)
    bonus = basic if i % 3 == 0 else 0
    overtime = random.randint(0, 3000)
    days_present = random.randint(22, 26)
    pf_emp = round(basic * 0.12)
    pf_empr = round(basic * 0.12)
    esi_emp = round((basic + hra) * 0.0075) if (basic + hra) <= 21000 else 0
    esi_empr = round((basic + hra) * 0.0325) if (basic + hra) <= 21000 else 0
    pt = 200
    tds = round(basic * 0.05) if basic > 40000 else 0
    loan = random.choice([0, 0, 0, 2000, 3000, 5000])
    advance = random.choice([0, 0, 1000, 2000])
    leave_ded = max(0, (26 - days_present) * round(basic / 26))
    gratuity = round(basic * 4.81 / 100)
    doj_year = random.randint(2015, 2023)
    doj_month = random.randint(1, 12)

    row = [
        emp_id, name, dept, desig, "March", 2025,
        26, days_present, basic, hra, conveyance,
        medical, special, bonus, overtime,
        pf_emp, pf_empr, esi_emp, esi_empr,
        pt, tds, loan, advance,
        leave_ded, gratuity,
        f"SBIN{random.randint(10000000000, 99999999999)}", f"SBIN000{random.randint(1000, 9999)}",
        f"ABCDE{random.randint(1000, 9999)}F", f"10010{random.randint(10000000000, 99999999999)}",
        f"{doj_year}-{doj_month:02d}-01"
    ]
    ws.append(row)

wb.save("/home/claude/payroll_system/employee_payroll_data.xlsx")
print("Excel created with 30 employees")
